"""
build_paro.py — Actualiza public/municipios_ine.json con datos del SEPE

USO: python scripts/build_paro.py

FICHERO NECESARIO en data/:
  PARO_MUNICIPIOS_MMMAA.xlsx — descargable desde:
  https://www.sepe.es/HomeSepe/que-es-el-sepe/estadisticas/datos-estadisticos/paro/datos-municipios.html

ESTRUCTURA DEL EXCEL:
  Hojas "PARO {PROVINCIA}":     paro registrado por municipio y sexo/edad/sector
  Hojas "CONTRATOS {PROVINCIA}": contratos registrados por municipio y tipo/sector

NOTA '<5': el SEPE censura valores entre 1 y 4 → se almacenan como 2
"""

import sys, json, glob
from pathlib import Path
try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

import warnings; warnings.filterwarnings('ignore')

ROOT = Path(__file__).parent.parent
DATA = ROOT / 'data'
OUT  = ROOT / 'public' / 'municipios_ine.json'

def to_int(v):
    if v is None: return None
    s = str(v).strip()
    if s in ('', ' '): return None
    if s == '<5': return 2
    try: return int(float(s))
    except: return None

def parse_paro(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    paro = {}; contratos = {}

    for sname in wb.sheetnames:
        if not (sname.startswith('PARO ') or sname.startswith('CONTRATOS ') or sname.startswith('CONTRTOS ')):
            continue
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        is_paro = sname.startswith('PARO ')
        for row in rows[8:]:
            if row[0] is None: continue
            try: cod = str(int(row[0])).zfill(5)
            except: continue
            if len(cod) != 5 or not cod.isdigit(): continue
            if is_paro:
                paro[cod] = {
                    'paro_total': to_int(row[2]),
                    'paro_h_m25': to_int(row[3]), 'paro_h_25_44': to_int(row[4]), 'paro_h_m45': to_int(row[5]),
                    'paro_m_m25': to_int(row[6]), 'paro_m_25_44': to_int(row[7]), 'paro_m_m45': to_int(row[8]),
                    'paro_agri': to_int(row[9]),  'paro_indus': to_int(row[10]),
                    'paro_constr': to_int(row[11]),'paro_serv': to_int(row[12]), 'paro_sin_emp': to_int(row[13]),
                }
            else:
                contratos[cod] = {
                    'contr_total': to_int(row[2]),
                    'contr_h_indef': to_int(row[3]),'contr_h_temp': to_int(row[4]),
                    'contr_m_indef': to_int(row[6]),'contr_m_temp': to_int(row[7]),
                    'contr_agri': to_int(row[9]),  'contr_indus': to_int(row[10]),
                    'contr_constr': to_int(row[11]),'contr_serv': to_int(row[12]),
                }
    wb.close()
    return paro, contratos

def main():
    # Find most recent PARO xlsx
    files = sorted(DATA.glob('PARO_MUNICIPIOS_*.xlsx'))
    if not files:
        sys.exit("No se encontró PARO_MUNICIPIOS_*.xlsx en data/")
    xlsx = files[-1]
    mes_anio = xlsx.stem.replace('PARO_MUNICIPIOS_', '')
    print(f"Procesando: {xlsx.name}")

    paro, contratos = parse_paro(xlsx)
    print(f"  Municipios con paro:      {len(paro):,}")
    print(f"  Municipios con contratos: {len(contratos):,}")

    print(f"\nCargando {OUT}...")
    with open(OUT) as f:
        municipios = json.load(f)

    for m in municipios:
        cod = m['cod']
        p = paro.get(cod, {})
        c = contratos.get(cod, {})
        m.update(p); m.update(c)
        m['anio_paro'] = mes_anio.replace('_',' ').title()
        pob_lab = m.get('pob_16_64')
        pt = m.get('paro_total')
        m['tasa_paro_est'] = round(pt/pob_lab*100, 1) if pt and pob_lab and pob_lab > 0 else None

    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(municipios, f, ensure_ascii=False, separators=(',', ':'))

    rurales = [m for m in municipios if m.get('densidad') and m['pob2025'] < 30000 and m['densidad'] < 100]
    paro_r = sum(m.get('paro_total') or 0 for m in rurales)
    lab_r  = sum(m.get('pob_16_64') or 0 for m in rurales)
    print(f"\n✓ {OUT} actualizado")
    print(f"  Paro registrado zona rural: {paro_r:,}")
    print(f"  Tasa paro est. rural:       {paro_r/lab_r*100:.1f}%" if lab_r else "")

if __name__ == '__main__': main()
