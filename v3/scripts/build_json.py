"""
build_json.py — Genera public/municipios_ine.json desde los ficheros del INE
Ejecutar: python scripts/build_json.py

FICHEROS NECESARIOS en data/:
  pobmun.zip              → https://www.ine.es/pob_xls/pobmun.zip
  Nacional_AAAA.zip       → INEbase > Nomenclátor > Fichero nacional ZIP (año más reciente)
  esp_municipios_*.csv    → REL, Ministerio de Política Territorial (superficie)
"""
import sys, json, csv, re, zipfile, io
from pathlib import Path
try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

ROOT  = Path(__file__).parent.parent
DATA  = ROOT / 'data'
OUT   = ROOT / 'public' / 'municipios_ine.json'

PROV_CCAA = {
    '01':'País Vasco','02':'Castilla-La Mancha','03':'Comunitat Valenciana',
    '04':'Andalucía','05':'Castilla y León','06':'Extremadura',
    '07':'Illes Balears','08':'Cataluña','09':'Castilla y León',
    '10':'Extremadura','11':'Andalucía','12':'Comunitat Valenciana',
    '13':'Castilla-La Mancha','14':'Andalucía','15':'Galicia',
    '16':'Castilla-La Mancha','17':'Cataluña','18':'Andalucía',
    '19':'Castilla-La Mancha','20':'País Vasco','21':'Andalucía',
    '22':'Aragón','23':'Andalucía','24':'Castilla y León',
    '25':'Cataluña','26':'La Rioja','27':'Galicia','28':'Comunidad de Madrid',
    '29':'Andalucía','30':'Región de Murcia','31':'Navarra',
    '32':'Galicia','33':'Asturias','34':'Castilla y León',
    '35':'Canarias','36':'Galicia','37':'Castilla y León',
    '38':'Canarias','39':'Cantabria','40':'Castilla y León',
    '41':'Andalucía','42':'Castilla y León','43':'Cataluña',
    '44':'Aragón','45':'Castilla-La Mancha','46':'Comunitat Valenciana',
    '47':'Castilla y León','48':'País Vasco','49':'Castilla y León',
    '50':'Aragón','51':'Ceuta','52':'Melilla'
}

def load_pobmun(data_dir):
    zips = sorted(data_dir.glob('pobmun*.zip'))
    if zips:
        zf = zipfile.ZipFile(zips[-1])
        newest = sorted([f for f in zf.namelist() if 'pobmun' in f and f.endswith('.xlsx')])[-1]
        wb = openpyxl.load_workbook(io.BytesIO(zf.read(newest)), read_only=True, data_only=True)
    else:
        xlsx = sorted(data_dir.glob('pobmun*.xlsx'))
        if not xlsx: raise FileNotFoundError("pobmun.zip o pobmun*.xlsx no encontrado en data/")
        wb = openpyxl.load_workbook(xlsx[-1], read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = rows[1]
    pob_col = next(i for i,h in enumerate(header) if h and str(h).startswith('POB'))
    anio = '20' + str(header[pob_col])[3:]
    result = {}
    for r in rows[2:]:
        if not r[0] or not r[2]: continue
        cod = str(r[0]).zfill(2) + str(r[2]).zfill(3)
        result[cod] = {'pob': int(r[pob_col]) if r[pob_col] else 0,
                       'nombre': str(r[3]).strip(), 'provincia': str(r[1]).strip(), 'anio': anio}
    print(f"  pobmun: {len(result):,} municipios · año {anio}")
    return result

def load_superficie(data_dir):
    csvs = sorted(data_dir.glob('esp_municipios*.csv')) + sorted(data_dir.glob('municipios*.csv'))
    if not csvs: raise FileNotFoundError("esp_municipios*.csv no encontrado en data/")
    result = {}
    with open(csvs[-1], encoding='utf-8', errors='replace') as f:
        for r in csv.reader(f):
            num = r[4].strip() if len(r) > 8 else ''
            if len(num) < 6: continue
            s = num[1:]; cod = s[:2]+s[2:5]
            try: result[cod] = float(r[8].strip().replace(',','.'))
            except: pass
    print(f"  superficie: {len(result):,} municipios")
    return result

def load_edad_nomenclator(data_dir):
    zips = sorted(data_dir.glob('Nacional_*.zip')) + sorted(data_dir.glob('nomdef*.zip'))
    result = {}; anio = '?'
    for zpath in zips:
        zf = zipfile.ZipFile(zpath)
        xlsx_files = [f for f in zf.namelist() if f.endswith('.xlsx')]
        if not xlsx_files: continue
        wb = openpyxl.load_workbook(io.BytesIO(zf.read(xlsx_files[0])), read_only=True, data_only=True)
        if 'Edad' not in wb.sheetnames: continue
        ws = wb['Edad']
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        # Extract year from header
        header = rows[0]
        anio = str(next((h for h in header if isinstance(h,int) or (isinstance(h,str) and '20' in h)), '?'))
        for r in rows[1:]:
            if not r[2] or not str(r[2]).startswith('000000'): continue
            cod = str(r[0]).zfill(2)+str(r[1]).zfill(3)
            result[cod] = {'p0_15': r[4], 'p16_64': r[5], 'p65': r[6]}
        print(f"  edad Nomenclátor: {len(result):,} municipios · año {anio}")
        break
    if not result:
        print("  edad: no se encontró fichero Nomenclátor ZIP en data/")
    return result, anio

def main():
    import warnings; warnings.filterwarnings('ignore')
    DATA.mkdir(exist_ok=True)
    print("Cargando fuentes...")
    pob = load_pobmun(DATA)
    sup = load_superficie(DATA)
    edad, anio_edad = load_edad_nomenclator(DATA)

    output = []
    for cod, p in sorted(pob.items()):
        s = sup.get(cod)
        e = edad.get(cod, {})
        pob2025 = p['pob']
        dens = round(pob2025/s, 2) if s and s > 0 else None
        p1664 = e.get('p16_64')
        p65 = e.get('p65')
        output.append({
            'cod': cod, 'nombre': p['nombre'], 'cpro': cod[:2],
            'provincia': p['provincia'], 'ccaa': PROV_CCAA.get(cod[:2],''),
            'pob2025': pob2025, 'sup_km2': s, 'densidad': dens,
            'pob_0_15': e.get('p0_15'), 'pob_16_64': p1664, 'pob_65_mas': p65,
            'pct_16_64': round(p1664/pob2025*100,1) if p1664 and pob2025 else None,
            'pct_65_mas': round(p65/pob2025*100,1) if p65 and pob2025 else None,
            'anio_pob': p['anio'], 'anio_edad': anio_edad, 'anio_sup': '2020',
        })

    OUT.parent.mkdir(exist_ok=True)
    with open(OUT,'w',encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, separators=(',',':'))
    rurales = sum(1 for m in output if m['densidad'] and m['pob2025']<30000 and m['densidad']<100)
    print(f"\n✓ {OUT}  ({len(output):,} municipios, {OUT.stat().st_size//1024} KB)")
    print(f"  Zonas rurales: {rurales:,} ({rurales/len(output)*100:.1f}%)")

if __name__ == '__main__': main()
